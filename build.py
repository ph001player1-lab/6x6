#!/usr/bin/env python3
"""
Сводит две старые базы в одну и печатает SQL для Supabase.

    python3 migrate/build.py Earth_Genesis.xlsx migrate/love_lottery.csv > migrate/import.sql

Правила слияния:
  • В Genesis один человек может занимать несколько строк — он менял город.
    Берём его последнюю запись по времени.
  • Если человек есть в обеих базах, побеждают более свежие ответы.
  • Ответы Genesis приходят названиями вариантов, лотереи — шестью цифрами.
    И то и другое сводится к c1…c6.
"""
import csv, sys, re
from datetime import datetime, timedelta

AXES = [
    ('Климат', ['Тропический','Субтропический','Умеренный','Холодный','Полярный']),
    ('Отношения', ['Моногамия','Полиамория','Свободные отношения','Асексуальный образ жизни']),
    ('Быт', ['Одиночное проживание','Патриархальная семья','Матриархальная семья',
             'Партнёрская семья','Коммуна','Гостевой брак']),
    ('Общество', ['Либерализм','Социализм','Консерватизм','Коммунизм','Анархизм','Авторитаризм']),
    ('Деньги', ['Предприниматель','Наёмный работник','Инвестор','Фрилансер',
                'Кооперативная экономика','Безусловный базовый доход']),
    ('Мировоззрение', ['Христианство','Ислам','Буддизм','Иудаизм','Научный материализм',
                       'Агностицизм','Эзотерика']),
]
RADIX = [5, 4, 6, 6, 6, 7]

def norm(s):
    return re.sub(r'\s+', ' ', str(s or '')).replace('ё', 'е').strip().lower()

LOOKUP = [{norm(lbl): i + 1 for i, lbl in enumerate(opts)} for _, opts in AXES]

def excel_date(v):
    if isinstance(v, datetime):
        return v
    return datetime(1899, 12, 30) + timedelta(days=float(v))

def read_genesis(path):
    from openpyxl import load_workbook
    ws = load_workbook(path, read_only=True)['Реестр городов']
    rows = list(ws.iter_rows(values_only=True))[1:]
    out = {}
    for r in rows:
        if not r or not r[2]:
            continue
        tid = int(float(r[2]))
        when = excel_date(r[0])
        try:
            answers = [LOOKUP[i][norm(r[6 + i])] for i in range(6)]
        except KeyError as e:
            print(f'-- пропущена строка {tid}: неизвестный вариант {e}', file=sys.stderr)
            continue
        rec = {
            'tg_id': tid,
            'username': str(r[1] or '').lstrip('@') or None,
            'first_name': str(r[3] or '').strip(),
            'answers': answers,
            'when': when,
            'src': 'genesis',
        }
        if tid not in out or when > out[tid]['when']:
            out[tid] = rec
    return out

def read_lottery(path):
    out = {}
    with open(path, encoding='utf-8') as f:
        for row in csv.DictReader(f):
            digits = row['answers'].strip()
            if not re.fullmatch(r'[1-7]{6}', digits):
                print(f"-- пропущен {row['uid']}: ответы «{digits}»", file=sys.stderr)
                continue
            answers = [int(d) for d in digits]
            if any(a > RADIX[i] for i, a in enumerate(answers)):
                print(f"-- пропущен {row['uid']}: ответ вне диапазона {answers}", file=sys.stderr)
                continue
            tid = int(row['uid'])
            out[tid] = {
                'tg_id': tid,
                'username': (row['nick'] or '').lstrip('@') or None,
                'first_name': (row['name'] or '').strip(),
                'answers': answers,
                'when': datetime.fromisoformat(row['updated']),
                'src': 'lottery',
            }
    return out

def q(v):
    return 'null' if v in (None, '') else "'" + str(v).replace("'", "''") + "'"

def main():
    genesis = read_genesis(sys.argv[1])
    lottery = read_lottery(sys.argv[2])

    merged, clashes = dict(genesis), []
    for tid, rec in lottery.items():
        old = merged.get(tid)
        if old:
            clashes.append((tid, old, rec))
            if rec['when'] <= old['when']:
                continue
        merged[tid] = rec

    print('-- Импорт объединённой базы 6×6.')
    print(f'-- Genesis: {len(genesis)} человек, лотерея: {len(lottery)}, '
          f'пересечение: {len(clashes)}, итого: {len(merged)}')
    for tid, old, new in clashes:
        same = 'ответы совпали' if old['answers'] == new['answers'] else \
               f"{old['answers']} → {new['answers']}"
        print(f'--   {tid} @{new["username"]}: {same}')
    print()
    print('begin;')
    for r in sorted(merged.values(), key=lambda x: x['when']):
        a = r['answers']
        print(
            'insert into players (tg_id, username, first_name, c1,c2,c3,c4,c5,c6,'
            ' created_at, answered_at) values ('
            f"{r['tg_id']}, {q(r['username'])}, {q(r['first_name'])}, "
            f"{a[0]},{a[1]},{a[2]},{a[3]},{a[4]},{a[5]}, "
            f"'{r['when'].isoformat()}', '{r['when'].isoformat()}')\n"
            '  on conflict (tg_id) do update set'
            ' username = excluded.username, first_name = excluded.first_name,'
            ' c1=excluded.c1, c2=excluded.c2, c3=excluded.c3,'
            ' c4=excluded.c4, c5=excluded.c5, c6=excluded.c6,'
            ' answered_at = excluded.answered_at'
            '  where players.answered_at < excluded.answered_at;'
        )
    print()
    print('-- Импорт не должен разослать пуши всем сразу:')
    print('delete from outbox where sent_at is null;')
    print('select refresh_axis_stats();')
    print('commit;')

if __name__ == '__main__':
    main()
